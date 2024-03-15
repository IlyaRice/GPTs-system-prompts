import zipfile
import os
import re
import json
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import NamedStyle, NumberFormatDescriptor, Font

# Function to extract the sender's name
def extract_sender_name(text_entities):
    for entity in text_entities:
        if entity.get("type") == "mention":
            return entity.get("text", "Unknown")
        elif entity.get("type") == "mention_name":
            name = entity.get("text", "Unknown")
            user_id = entity.get("user_id", "")
            return f"{name} (id{user_id})"
    return "Unknown"

# Function to determine the type and amount of payment
def extract_payment_info(text, patterns):
    for payment_type, pattern in patterns.items():
        match = re.search(pattern, text)
        if match:
            # Changing the format of the amount for correct display in Excel
            amount = float(match.group(1).replace(' ', '').replace(',', '.').rstrip('₽€'))
            return payment_type, amount
    return None, None

# Pre-compiling regular expressions
patterns = {
    "New subscription": re.compile(r"оформил подписку на ваш канал «.*» за (\d{1,3}(?: \d{3})*\.\d{2}[₽€])"),
    "Subscription renewal": re.compile(r"Поступил еще один платёж в размере (\d{1,3}(?: \d{3})*\.\d{2}[₽€])"),
    "Donation": re.compile(r"Вы получили (\d{1,3}(?: \d{3})*\.\d{2}[₽€])")
}

def process_data(json_file_path):
    # Loading a new JSON file
    with open(json_file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)

    # Data processing
    result = []
    for message in data["messages"]:
        if message.get("from") != "Donate":
            continue

        sender_name = extract_sender_name(message.get("text_entities", []))
        # Converting the date string into a datetime object
        message_date = datetime.fromisoformat(message.get("date"))

        for entity in message.get("text_entities", []):
            if entity.get("type") == "plain":
                payment_type, amount = extract_payment_info(entity.get("text", ""), patterns)
                if payment_type and amount:
                    result.append({
                        "Date and time": message_date,
                        "User": sender_name,
                        "Amount": amount,
                        "Category": payment_type
                    })

    # Creating DataFrame
    df = pd.DataFrame(result)

    # Function to determine the quarter
    def get_quarter(month):
        if 1 <= month <= 3:
            return 'Q1'
        elif 4 <= month <= 6:
            return 'Q2'
        elif 7 <= month <= 9:
            return 'Q3'
        else:
            return 'Q4'

    # Adding a column with quarters
    df['Quarter'] = df['Date and time'].apply(lambda x: get_quarter(x.month))
    return df
 
def style_excel_sheet(worksheet, start_row):
    # Formatting settings
    text_format = '@'  # Text format
    date_style = NamedStyle(name='datetime', number_format='DD.MM.YYYY HH:MM')
    currency_style = NamedStyle(name='currency', number_format='# ##0.00 ₽')

    # Formatting the main table starting from start_row
    for row in worksheet.iter_rows(min_row=start_row, max_col=4, max_row=worksheet.max_row):
        # Formatting the date column
        row[0].style = date_style

        # Formatting the "User" and "Category" columns as text
        for cell in [row[1], row[3]]:
            cell.number_format = text_format

        # Formatting the amount column
        row[2].style = currency_style

    # Setting column widths
    worksheet.column_dimensions['A'].width = 18
    worksheet.column_dimensions['B'].width = 22
    worksheet.column_dimensions['C'].width = 10
    worksheet.column_dimensions['D'].width = 15

    # Setting text alignment to the left
    for col in worksheet.iter_cols(min_row=start_row, max_col=4, max_row=worksheet.max_row):
        for cell in col:
            cell.alignment = openpyxl.styles.Alignment(horizontal='left')

def calculate_summary(df):
    # Initializing a dictionary for summary information
    summary = {'Donation': [0, 0.0], 'New and renewed subscriptions': [0, 0.0], 'Subscription tax 6%': [0, 0.0], 'Total payments': [0, 0.0]}
    
    # Calculating the total number and amount for Donations
    if 'Donation' in df['Category'].values:
        donation_summary = df[df['Category'] == 'Donation']['Amount'].agg(['count', 'sum'])
        summary['Donation'] = [donation_summary['count'], donation_summary['sum']]

    # Calculating the total number and amount for New and Renewed subscriptions
    subscription_df = df[df['Category'].isin(['New subscription', 'Subscription renewal'])]
    if not subscription_df.empty:
        subscription_summary = subscription_df['Amount'].agg(['count', 'sum'])
        summary['New and renewed subscriptions'] = [subscription_summary['count'], subscription_summary['sum']]

        # Calculating the tax on subscriptions
        summary['Subscription tax 6%'] = [0, subscription_summary['sum'] * 0.06]

    # Total number and amount of payments
    summary['Total payments'] = [df.shape[0], df['Amount'].sum()]

    return summary

def add_summary_to_excel(worksheet, summary):
    # Inserting summary information before the main table
    for r_idx, (key, values) in enumerate(summary.items(), 1):
        # Inserting the category name
        worksheet.cell(row=r_idx, column=1, value=key)

        # Inserting the number and amount
        for c_idx, value in enumerate(values, 2):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = openpyxl.styles.Alignment(horizontal='left')
            if c_idx == 3:  # Formatting the amount column
                cell.number_format = '# ##0.00 ₽'

def save_to_excel(df):
    output_files = []
    quarter_summaries = {}
    quarter_to_months = {
        'Q1': '_jan_feb_mar',
        'Q2': '_apr_may_jun',
        'Q3': '_jul_aug_sep',
        'Q4': '_oct_nov_dec'
    }

    for quarter in ['Q1', 'Q2', 'Q3', 'Q4']:
        quarter_df = df[df['Quarter'] == quarter]
        if quarter_df.empty:
            quarter_summaries[quarter + quarter_to_months[quarter]] = "No data for quarter {}".format(quarter)
            continue

        quarter_df = quarter_df.drop(columns=['Quarter'])
        summary = calculate_summary(quarter_df)

        # Formatting summary data for return
        formatted_summary = '\n'.join([
            f"{key} - {value[0]}pcs, {value[1]:,.2f}₽" if key != 'Subscription tax 6%' else f"{key} - {value[1]:,.2f}₽"
            for key, value in summary.items()
        ])
        quarter_summaries[quarter + quarter_to_months[quarter]] = formatted_summary

        # Modified file name format
        file_name = f'/mnt/data/Quarter_{quarter}{quarter_to_months[quarter]}.xlsx'
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            workbook = writer.book
            worksheet = workbook.create_sheet(f'Quarter {quarter}{quarter_to_months[quarter]}')

            add_summary_to_excel(worksheet, summary)

            # Leaving one empty row between the summary and the main table
            df_start_row = worksheet.max_row + 2

            # Saving data
            for r_idx, row in enumerate(quarter_df.to_numpy(), df_start_row):
                for c_idx, value in enumerate(row, 1):
                    worksheet.cell(row=r_idx, column=c_idx, value=value)

            style_excel_sheet(worksheet, df_start_row)

        output_files.append(file_name)

    return output_files, quarter_summaries
